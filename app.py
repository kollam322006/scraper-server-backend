import os
import io
import json
import time
import requests
import re
import logging
from datetime import datetime
from typing import List
from urllib.parse import urlparse, urlunparse

from bs4 import BeautifulSoup

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import select, func, or_
from sqlalchemy.orm import relationship

# RQ/Redis for background tasks
import redis
from rq import Queue

# XLSX library imports
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- 1. Configuration and Initialization ---

# Set up logging for better debugging
logging.basicConfig(level=logging.INFO)
app = Flask(__name__)

# Environment variables for configuration
DB_URL = os.environ.get("DATABASE_URL", "sqlite:///scraper.db")
# Use a separate URL for Redis, default to localhost for development
REDIS_URL = os.environ.get("REDIS_URL", "redis://localhost:6379")

# Flask and SQLAlchemy Configuration
# Replace postgres:// with postgresql:// for compatibility with SQLAlchemy 2.0+
app.config["SQLALCHEMY_DATABASE_URI"] = DB_URL.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "a_very_secret_key")

# Initialize Extensions
db = SQLAlchemy(app)
CORS(app, resources={r"/*": {"origins": "*"}}) # Allow all origins

# Initialize Redis and RQ Queue
redis_conn = redis.from_url(REDIS_URL)
q = Queue(connection=redis_conn)

# --- 2. SQLAlchemy Models ---

# Association table for many-to-many relationship between TargetURL and Email
url_email_association = db.Table(
    "url_email_association",
    db.Column("target_url_id", db.Integer, db.ForeignKey("target_url.id"), primary_key=True),
    db.Column("email_id", db.Integer, db.ForeignKey("email.id"), primary_key=True),
)

class Batch(db.Model):
    __tablename__ = 'batch'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=True)
    job_id = db.Column(db.String(36), unique=True, nullable=True) # RQ Job ID
    status = db.Column(db.String(50), default="PENDING")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    start_time = db.Column(db.DateTime, nullable=True)
    end_time = db.Column(db.DateTime, nullable=True)

    target_urls = relationship("TargetURL", backref="batch", lazy=True, cascade="all, delete-orphan")
    
    # Relationship to find all unique emails associated with this batch
    # This relationship definition is complex but aims to get a list of unique Email objects
    emails = relationship(
        "Email", 
        secondary=url_email_association, 
        primaryjoin=lambda: Batch.id == TargetURL.batch_id,
        secondaryjoin=lambda: TargetURL.id == url_email_association.c.target_url_id,
        overlaps="target_urls,emails",
        viewonly=True 
    )


class TargetURL(db.Model):
    __tablename__ = 'target_url'
    id = db.Column(db.Integer, primary_key=True)
    url = db.Column(db.Text, nullable=False)
    domain = db.Column(db.String(255), nullable=False)
    status = db.Column(db.String(50), default="PENDING") # PENDING, IN_PROGRESS, COMPLETED, FAILED
    error_message = db.Column(db.Text, nullable=True)
    batch_id = db.Column(db.Integer, db.ForeignKey("batch.id"), nullable=False)

    emails = relationship("Email", secondary=url_email_association, backref="target_urls", lazy=True, cascade="save-update")


class Email(db.Model):
    __tablename__ = 'email'
    id = db.Column(db.Integer, primary_key=True)
    address = db.Column(db.String(255), unique=True, nullable=False, index=True)
    domain = db.Column(db.String(255), nullable=False, index=True)
    # The ID of the batch where this email was first discovered
    first_found_batch_id = db.Column(db.Integer, db.ForeignKey('batch.id'), nullable=True) 

# --- 3. Utility Functions for Scraping (Run by RQ Worker) ---

# Define common file extensions to ignore globally
FILE_EXTENSIONS = (
    '.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.ico',
    '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx',
    '.zip', '.rar', '.7z', '.mp4', '.avi', '.mov', '.mp3', '.wav',
    '.js', '.css', '.xml', '.json', '.txt', '.php', '.asp', '.html',
    '.exe', '.dll', '.tif', '.tiff', '.bmp', '.mpv', '.mov', '.flv'
)

def normalize_url(url: str) -> str:
    """Ensures the URL has a scheme (defaulting to https) and is cleaned."""
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    parsed = urlparse(url)
    # Ensure netloc is present and domain is lowercase
    if not parsed.netloc:
        return ""
    
    # Keep only scheme, netloc, and path
    return urlunparse(parsed._replace(params='', query='', fragment=''))

def extract_domain(url: str) -> str:
    """Extracts the root domain from a full URL."""
    try:
        parsed_uri = urlparse(url)
        domain = '{uri.netloc}'.format(uri=parsed_uri)
        # Handle cases where domain starts with 'www.'
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain.lower()
    except Exception:
        return ""

def generate_target_urls(raw_inputs: List[str]) -> List[tuple]:
    """
    Takes raw domain/URL inputs and generates normalized, target URLs 
    (contact, about, privacy, etc.)
    Returns a list of tuples: (normalized_url, domain)
    """
    base_urls = []
    
    # 1. Normalize all inputs to root domains or full URLs
    for input_str in raw_inputs:
        input_str = input_str.strip()
        if not input_str:
            continue
        
        normalized = normalize_url(input_str)
        if normalized:
            base_urls.append(normalized)

    # 2. Generate policy page URLs
    target_pages = set()
    suffixes = [
        "contact", "contact-us", "about", "about-us", 
        "policies/privacy-policy", "policies/contact-information"
    ]
    
    for base_url in base_urls:
        domain = extract_domain(base_url)
        if not domain:
            continue

        # If input was a domain/root, generate all policy URLs
        is_root = (base_url == f"https://{domain}" or base_url == f"https://www.{domain}")
        
        # Always add the root URL as a target page
        target_pages.add((f"https://{domain}", domain))
        
        if is_root:
            for suffix in suffixes:
                # Add / before suffix if needed
                page_url = f"{base_url.rstrip('/')}/{suffix.lstrip('/')}"
                target_pages.add((page_url, domain))
        else:
            # If input was a specific policy page, just add that one
            target_pages.add((base_url, domain))

    app.logger.debug(f"Generated {len(target_pages)} target URLs from {len(raw_inputs)} raw inputs.")
    return list(target_pages)


def scrape_task(batch_id: int, target_urls: List[tuple]):
    """
    Performs the background scraping task for a batch.
    """
    app.logger.info(f"Starting scrape task for Batch {batch_id} with {len(target_urls)} URLs.")
    
    try:
        # 1. Initialize Batch Status
        with app.app_context():
            # Accessing model classes (Batch, TargetURL, Email) from the module scope
            batch = db.session.get(Batch, batch_id) 
            if not batch:
                app.logger.error(f"Batch with ID {batch_id} not found.")
                return

            batch.status = "IN_PROGRESS"
            batch.start_time = datetime.now()
            db.session.commit()

        # 2. Iterate and scrape
        # Tracks unique emails found in this batch (address -> Email Object)
        unique_emails_in_batch = {} 

        for url_id, url_full, url_domain in target_urls:
            app.logger.info(f"Worker processing URL ID {url_id}: {url_full}")
            
            with app.app_context():
                target_url_obj = db.session.get(TargetURL, url_id)
                if not target_url_obj:
                    app.logger.error(f"TargetURL {url_id} not found.")
                    continue
                
                target_url_obj.status = "IN_PROGRESS"
                db.session.commit()
            
            try:
                # --- Fetching and Parsing ---
                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
                response = requests.get(url_full, headers=headers, timeout=15)
                response.raise_for_status() 
                page_content = response.text
                
                # --- Email Extraction and Filtering ---
                # Regex for potential emails
                potential_emails = set(re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', page_content))
                
                def get_email_domain(email):
                    return email.split('@')[-1]

                common_info_prefixes = ["info", "contact", "support", "sales"]
                expected_domain = url_domain.lower()

                filtered_emails = set()
                
                for email in potential_emails:
                    email_lower = email.lower()
                    email_domain = get_email_domain(email_lower)
                    email_prefix = email_lower.split('@')[0].lower()
                    
                    # File Extension Filter: Skip if it ends like a common file
                    is_file_extension = any(email_lower.endswith(ext) for ext in FILE_EXTENSIONS)
                    
                    if is_file_extension:
                        continue 

                    # Domain Relevance Filter: Keep if it matches the target domain or is a common info/contact prefix
                    if email_domain == expected_domain or email_prefix in common_info_prefixes:
                        filtered_emails.add(email_lower)
                
                app.logger.info(f"Filtered down to {len(filtered_emails)} unique relevant emails.")


                # 3. Save emails and link them to the TargetURL (Transactional)
                with app.app_context():
                    
                    emails_to_link = []

                    for email_address in filtered_emails:
                        
                        # Check local batch tracker first (avoids duplicate DB query per batch)
                        if email_address not in unique_emails_in_batch:
                            
                            # Check database globally
                            email_obj = db.session.scalar(
                                select(Email).filter_by(address=email_address)
                            )
                            
                            if not email_obj:
                                # Create a new Email object if it doesn't exist
                                email_obj = Email(
                                    address=email_address, 
                                    first_found_batch_id=batch_id,
                                    domain=get_email_domain(email_address)
                                )
                                db.session.add(email_obj)
                                db.session.flush() # Ensure object gets an ID before session commit
                                
                            # Add to batch tracker for future URLs in this batch
                            unique_emails_in_batch[email_address] = email_obj
                        
                        # Use the existing object (either fetched or created)
                        email_obj = unique_emails_in_batch[email_address]
                        emails_to_link.append(email_obj)

                    # Update TargetURL with status and results
                    target_url_obj = db.session.get(TargetURL, url_id)
                    # Use set() conversion to prevent duplicate links if the email was found multiple times on the same page
                    current_email_ids = {e.id for e in target_url_obj.emails}
                    new_emails_to_add = [e for e in emails_to_link if e.id not in current_email_ids]
                    
                    target_url_obj.status = "COMPLETED"
                    target_url_obj.emails.extend(new_emails_to_add) 
                    db.session.commit()


            except requests.exceptions.RequestException as e:
                app.logger.error(f"Request error for {url_full}: {e}")
                with app.app_context():
                    target_url_obj = db.session.get(TargetURL, url_id)
                    if target_url_obj:
                        target_url_obj.status = "FAILED"
                        target_url_obj.error_message = f"Request error: {str(e)[:250]}"
                        db.session.commit()
            
            except Exception as e:
                app.logger.error(f"General error during scraping {url_full}: {e}")
                with app.app_context():
                    target_url_obj = db.session.get(TargetURL, url_id)
                    if target_url_obj:
                        target_url_obj.status = "FAILED"
                        target_url_obj.error_message = f"Internal error: {str(e)[:250]}"
                        db.session.commit()


        # 4. Finalize Batch Status
        with app.app_context():
            # Check if any TargetURLs are still IN_PROGRESS (should be 0 if the loop finished)
            pending_count = db.session.scalar(
                select(func.count(TargetURL.id))
                .filter(TargetURL.batch_id == batch_id)
                .filter(TargetURL.status == "IN_PROGRESS")
            )

            batch = db.session.get(Batch, batch_id)
            batch.end_time = datetime.now()
            
            if pending_count == 0:
                batch.status = "COMPLETED"
                app.logger.info(f"Batch {batch_id} finished successfully.")
            else:
                # If there are still IN_PROGRESS items, it means an unhandled error occurred or the worker crashed mid-batch.
                # Since the worker finished its assigned list, we mark it based on success/failure in that list.
                # A more accurate check: did ALL TargetURLs fail?
                total_urls = db.session.scalar(select(func.count(TargetURL.id)).filter_by(batch_id=batch_id))
                completed_or_failed_urls = db.session.scalar(
                    select(func.count(TargetURL.id))
                    .filter(TargetURL.batch_id == batch_id)
                    .filter(or_(TargetURL.status == "COMPLETED", TargetURL.status == "FAILED"))
                )
                
                if completed_or_failed_urls == total_urls:
                    batch.status = "COMPLETED" # All finished, even if some failed.
                else:
                    batch.status = "FAILED" # Should not happen if the loop completes cleanly
                    app.logger.error(f"Batch {batch_id} finished with {total_urls - completed_or_failed_urls} unprocessed URLs.")

            db.session.commit()
            
    except Exception as e:
        app.logger.error(f"Unrecoverable error in scrape_task for batch {batch_id}: {e}")
        # Mark batch as FAILED if an exception was raised outside the inner loop
        with app.app_context():
            batch = db.session.get(Batch, batch_id)
            if batch:
                batch.status = "FAILED"
                db.session.commit()


# --- 4. API Routes ---

@app.route("/create_tables", methods=["POST"])
def create_tables():
    """Endpoint to ensure database tables are created."""
    with app.app_context():
        # Check for existing tables by querying a model
        try:
            # Check if the Batch table exists and is readable
            db.session.execute(select(Batch).limit(1))
            db_exists = True
        except Exception:
            db_exists = False

        if not db_exists:
            db.create_all()
            return jsonify({"message": "Tables created."}), 201
            
        return jsonify({"message": "Tables created or already exist."}), 200


@app.route("/start_batch_scrape", methods=["OPTIONS", "POST"])
def start_batch_scrape():
    """Initiates a new scraping batch."""
    if request.method == "OPTIONS":
        return "", 200

    data = request.get_json()
    raw_inputs = data.get("domains", "").split("\n")
    batch_name = data.get("name", f"Batch {datetime.now().strftime('%m/%d/%Y, %I:%M:%S %p')}")

    if not raw_inputs or all(not i.strip() for i in raw_inputs):
        return jsonify({"error": "No domains or URLs provided."}), 400

    target_urls_data = generate_target_urls(raw_inputs)
    
    if not target_urls_data:
        return jsonify({"error": "Could not generate valid URLs from input."}), 400

    with app.app_context():
        # 1. Create Batch record
        new_batch = Batch(name=batch_name, status="PENDING")
        db.session.add(new_batch)
        db.session.flush() # Get the batch ID before commit

        # 2. Create TargetURL records
        target_url_tuples = []
        for url, domain in target_urls_data:
            target_url = TargetURL(url=url, domain=domain, batch_id=new_batch.id)
            db.session.add(target_url)
            db.session.flush() # Get the URL ID before commit
            # Store (id, full_url, domain) for the worker task
            target_url_tuples.append((target_url.id, url, domain))

        db.session.commit()

        # 3. Enqueue job
        # RQ only takes simple types, so we pass the TargetURL data needed for the worker
        job = q.enqueue(scrape_task, new_batch.id, target_url_tuples, job_timeout='3h')
        new_batch.job_id = job.id
        db.session.commit()

    return jsonify({
        "message": "Scraping job started",
        "batch_id": new_batch.id,
        "job_id": job.id,
        "total_urls": len(target_url_tuples)
    }), 202


@app.route("/status/<job_id>", methods=["GET"])
def get_job_status(job_id):
    """Retrieves the status of an RQ job."""
    try:
        job = q.fetch_job(job_id)
        if job is None:
            # If the job is missing from RQ, check DB for final status
            with app.app_context():
                batch = db.session.scalar(select(Batch).filter_by(job_id=job_id))
                if batch:
                    return jsonify({"job_id": job_id, "status": batch.status, "progress": 100}), 200
            
            return jsonify({"error": "Job not found"}), 404
        
        # Calculate progress based on DB records (more accurate than RQ metadata)
        with app.app_context():
            batch = db.session.scalar(select(Batch).filter_by(job_id=job_id))
            if not batch:
                return jsonify({"job_id": job_id, "status": job.get_status(), "progress": 0}), 200

            total_urls = db.session.scalar(select(func.count(TargetURL.id)).filter_by(batch_id=batch.id))
            processed_urls = db.session.scalar(
                select(func.count(TargetURL.id))
                .filter(TargetURL.batch_id == batch.id)
                .filter(or_(
                    TargetURL.status == "COMPLETED",
                    TargetURL.status == "FAILED"
                ))
            )
            
            progress = (processed_urls / total_urls) * 100 if total_urls else 0
            
            return jsonify({
                "job_id": job_id, 
                "status": batch.status, 
                "progress": round(progress),
                "processed_count": processed_urls,
                "total_count": total_urls
            }), 200

    except Exception as e:
        app.logger.error(f"Error fetching job status: {e}")
        return jsonify({"error": "Failed to retrieve job status."}), 500


@app.route("/batches", methods=["GET"])
def get_batches():
    """Retrieves a list of all existing batches."""
    with app.app_context():
        batches = db.session.scalars(select(Batch).order_by(Batch.created_at.desc())).all()
        
        batch_list = []
        for batch in batches:
            # FIX: Count unique email IDs linked to this batch.
            email_count = db.session.scalar(
                select(func.count(func.distinct(Email.id)))
                .join(url_email_association)
                .join(TargetURL)
                .filter(TargetURL.batch_id == batch.id)
            )
            
            batch_list.append({
                "id": batch.id,
                "name": batch.name,
                "status": batch.status,
                "job_id": batch.job_id,
                "created_at": batch.created_at.isoformat(),
                "emails_count": email_count if email_count is not None else 0
            })
        return jsonify(batch_list), 200


@app.route("/batches/<int:batch_id>/emails", methods=["GET"])
def get_batch_emails(batch_id):
    """Retrieves all unique emails associated with a specific batch."""
    with app.app_context():
        # Query unique emails that are associated with any TargetURL belonging to this batch
        emails_query = select(Email).distinct().join(Email.target_urls).filter(TargetURL.batch_id == batch_id)
        emails = db.session.scalars(emails_query).all()

        emails_list = [
            {
                "id": email.id,
                "address": email.address,
                "domain": email.domain,
                "source_urls": [url.url for url in email.target_urls if url.batch_id == batch_id]
            }
            for email in emails
        ]
        return jsonify(emails_list), 200

@app.route("/batches/<int:batch_id>/urls", methods=["GET"])
def get_batch_urls(batch_id):
    """Retrieves all target URLs for a batch and their statuses."""
    with app.app_context():
        urls = db.session.scalars(select(TargetURL).filter_by(batch_id=batch_id)).all()
        url_list = [{
            "id": url.id,
            "url": url.url,
            "status": url.status,
            "error_message": url.error_message,
            "emails_found": len(url.emails)
        } for url in urls]
        return jsonify(url_list), 200


@app.route("/batches/<int:batch_id>/export_xlsx", methods=["GET"])
def export_xlsx(batch_id):
    """Generates and serves an Excel file for the unique emails in a batch."""
    with app.app_context():
        # Use the same query logic as /batches/<int:batch_id>/emails to fetch the data
        emails_query = select(Email).distinct().join(Email.target_urls).filter(TargetURL.batch_id == batch_id)
        emails = db.session.scalars(emails_query).all()
        
        if not emails:
            return jsonify({"message": "No emails found for this batch."}), 404

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Emails"

        # Headers
        headers = ["Email Address", "Domain", "Source URLs in Batch"]
        ws.append(headers)

        # Data rows
        for email in emails:
            # Filter source URLs to only include those associated with the current batch
            source_urls = [
                url.url for url in email.target_urls 
                if url.batch_id == batch_id
            ]
            ws.append([
                email.address,
                email.domain,
                ", ".join(source_urls) 
            ])

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30

        # Save to a virtual file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        batch = db.session.get(Batch, batch_id)
        # Safely create filename
        batch_name_safe = batch.name.replace(' ', '_').replace('/', '-').replace('\\', '-') if batch else "unknown_batch"
        filename = f"emails_batch_{batch_id}_{batch_name_safe}.xlsx"

        return send_file(
            output, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )


@app.route("/batches/delete", methods=["OPTIONS", "DELETE"])
def delete_batches():
    """Deletes batches and their related records."""
    if request.method == "OPTIONS":
        return "", 200
        
    data = request.get_json()
    batch_ids = data.get("batch_ids", [])
    
    if not batch_ids:
        return jsonify({"message": "No batch IDs provided."}), 400

    with app.app_context():
        try:
            # 1. Delete TargetURL records associated with the batches
            # SQLAlchemy's cascade on TargetURL should handle the association table entries
            target_urls_to_delete = db.session.scalars(
                select(TargetURL).filter(TargetURL.batch_id.in_(batch_ids))
            ).all()
            for url in target_urls_to_delete:
                db.session.delete(url)
            
            # 2. Delete the Batch records
            batches_to_delete = db.session.scalars(
                select(Batch).filter(Batch.id.in_(batch_ids))
            ).all()
            
            deleted_count = len(batches_to_delete)
            for batch in batches_to_delete:
                db.session.delete(batch)
            
            # 3. Handle orphaned Email records (optional, depending on business logic, but safe to ignore here)
            # The current logic keeps Email records even if their associated TargetURLs are deleted.
            # They are only deleted if all TargetURLs that found them are gone.

            db.session.commit()
            
            return jsonify({"message": f"Successfully deleted {deleted_count} batches and their related data."}), 200
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Failed to delete batches: {e}")
            return jsonify({"error": f"Failed to delete batches: {str(e)}"}), 500


# --- 5. App Runner ---

if __name__ == "__main__":
    # Create tables on application start (can be moved to a separate init script)
    with app.app_context():
        db.create_all()
    
    # Running this script directly will only start the Flask web server.
    # The RQ worker process must be started separately (e.g., using a separate `rq worker` command).
    app.run(host='0.0.0.0', port=os.environ.get('PORT', 5000))
